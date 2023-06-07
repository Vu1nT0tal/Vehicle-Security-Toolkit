#!/usr/bin/python3

import re
import os
import sys
import json
import logging
import argparse
import subprocess
import configparser
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from rich.logging import RichHandler
from rich.pretty import pretty_repr


def setup_log():
    logging.getLogger("pygdbmi.GdbController").setLevel(logging.DEBUG)
    logging.getLogger("pygdbmi.IoManager").setLevel(logging.INFO)
    logging.basicConfig(
        level="NOTSET",
        format="%(message)s",
        datefmt="[%X]",
        handlers=[RichHandler(rich_tracebacks=True, tracebacks_show_locals=True, highlighter=None)],
    )


class RuntimeError(Exception):
    pass


class RunCmdFailedError(RuntimeError):
    pass


class GitRepository:
    def __init__(self, name: str, folder: str, gerrit_usr: str = None) -> None:
        self.__setup(name, folder, gerrit_usr)

    def __setup(self, name: str, folder: str, gerrit_usr: str = None) -> None:
        self.__setup_env()
        self.name = name
        self.folder = folder
        self.branch = None
        self.remote_branch = None
        self.head_commit = None
        self.local_branch_list = []
        self.remote_branch_list = []
        self.branch_list = []
        self.remote_list = None
        self.git_commits = []
        self.gerrit_usr = gerrit_usr

        self.__get_git_version()

        self.get_current_branch()
        self.get_remote_branch()
        self.get_current_head_commit()
        self.get_all_branches()
        self.get_remote_list()

    def __setup_env(self):
        os.environ['LANG'] = "en_US.UTF-8"
        os.environ['LANGUAGE'] = "en_US:en"

    def __run_cmd(self, cmd: str, cwd: str = None):
        command = cmd
        output = str()
        error = str()
        old_cwd = os.getcwd()
        if cwd:
            os.chdir(cwd)
        proc = subprocess.Popen(command, shell=True,
                                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        for line in proc.stdout:
            output += line.decode('utf-8')
        for line in proc.stderr:
            error += line.decode('utf-8')
        proc.communicate()
        proc.wait()
        proc.stdout.close()
        proc.stderr.close()
        if cwd:
            os.chdir(old_cwd)
        if proc.returncode != 0:
            raise RunCmdFailedError("run cmd \"%s\" failed, ret %d, out: %s, err: %s" % (
                command, proc.returncode, output, error))
        return (proc.returncode, output, error)

    def __get_git_version(self) -> str:
        ret, result, err = self.__run_cmd("git --version")
        if ret:
            logging.error(
                "failed to get git version, ret %d, out: %s, err: %s" % (ret, result, err))
            return None
        self.git_version = result.strip()
        logging.info(f"git version is: {self.git_version}")

    def get_current_branch(self) -> str:
        cmd = "git rev-parse --abbrev-ref --symbolic-full-name HEAD"
        ret, result, err = self.__run_cmd(cmd, self.folder)
        if ret != 0:
            logging.error("failed to get current branch for folder %s, ret %d, out: %s, err: %s" % (
                self.folder, ret, result, err))
            return None
        self.branch = result.strip()
        logging.debug(f"folder {self.folder} current branch: {self.branch}")
        return self.branch

    def get_remote_branch(self) -> str:
        cmd = "git rev-parse --abbrev-ref --symbolic-full-name @{upstream}"
        try:
            ret, result, err = self.__run_cmd(cmd, self.folder)
            if not ret:
                remote_branch = result.strip()
        except RunCmdFailedError as e:
            if "HEAD does not point to a branch" not in e.args[0]:
                raise e
            ret, result, err = self.__run_cmd("git show -s --pretty=%D HEAD", self.folder)
            remote_branch = "UNKNOWN" if ret else result.strip("HEAD, ").strip()
        self.remote_branch = remote_branch.strip()
        logging.debug(f"folder {self.folder} remote branch: {self.remote_branch}")
        return self.remote_branch

    def get_current_head_commit(self) -> str:
        cmd = "git rev-parse HEAD"
        ret, result, err = self.__run_cmd(cmd, self.folder)
        if ret != 0:
            logging.error("failed to get current commit for folder %s, ret %d, out: %s, err: %s" % (
                self.folder, ret, result, err))
            return None
        self.head_commit = result.strip()
        logging.debug(f"folder {self.folder} current HEAD commit: {self.head_commit}")
        return self.head_commit

    def get_remote_branch_list(self) -> list:
        cmd = "git branch -r | grep -v '\->'"
        ret, result, err = self.__run_cmd(cmd, self.folder)
        if ret != 0:
            logging.error("failed to get remote branch for folder %s, ret %d, out: %s, err: %s" % (
                self.folder, ret, result, err))
            return None
        self.remote_branch_list = []
        self.remote_branch_list.extend(
            branch.strip() for branch in result.splitlines()
        )
        return self.remote_branch_list

    def get_local_branch_list(self) -> list:
        cmd = "git branch -l"
        ret, result, err = self.__run_cmd(cmd, self.folder)
        if ret != 0:
            logging.error("failed to get local branch for folder %s, ret %d, out: %s, err: %s" % (
                self.folder, ret, result, err))
            return None
        self.local_branch_list = []
        self.local_branch_list.extend(branch.strip() for branch in result.splitlines())
        return self.local_branch_list

    def get_remote_list(self) -> list:
        ret, result, err = self.__run_cmd("git remote", self.folder)
        if ret != 0:
            logging.error("failed to get remote for folder %s, ret %d, out: %s, err: %s" % (
                self.folder, ret, result, err))
            return None
        self.remote_list = []
        self.remote_list.extend(branch.strip() for branch in result.splitlines())
        return self.remote_list

    def fetch_all_remotes(self) -> None:
        ret, result, err = self.__run_cmd("git fetch --all", self.folder)
        if ret != 0:
            logging.error("failed to get fetch all remote for folder %s, ret %d, out: %s, err: %s" % (
                self.folder, ret, result, err))
            return None
        self.get_all_branches()

    def __get_change_id_from_msg_for_commit(self, commit_id: str) -> str:
        ret, result, err = self.__run_cmd(f"git log {commit_id} -1", self.folder
        )
        if ret != 0:
            logging.error("failed to get commit msg for commit id %s, ret %d, out: %s, err: %s" % (
                commit_id, ret, result, err))
            return None
        CHANGE_ID_RE = re.compile(
            r'([\s]{0,})Change-Id:([\s]{0,})(?P<change_id>[0-9A-Za-z]{7,})')
        change_id = str()
        for line in result.splitlines():
            if match := CHANGE_ID_RE.match(line):
                change_id = match["change_id"].strip()
        return change_id

    def __get_git_commit_from_compare_result(self, result: str) -> list:
        commits = []
        GIT_MSG_RE = re.compile(r'(?P<commit_id>[0-9A-Fa-f]{7,})'
                                r'([\s]{1,})-([\s]{1,})'
                                r'(?P<commit_msg>[\u4e00-\u9fa5\w\W^\(^\)]+)'
                                r'\((?P<commit_time>[\w\W+]{1,})\)([\s]{1,})'
                                r'<(?P<commit_author>[\w\W+]{1,})>')
        for line in result.splitlines():
            if match := GIT_MSG_RE.match(line):
                commit_id = match["commit_id"].strip()
                change_id = self.__get_change_id_from_msg_for_commit(commit_id)
                commit = {
                    "commit_id": commit_id,
                    "change_id": change_id,
                    "commit_msg": match["commit_msg"].strip(),
                    "commit_time": match["commit_time"].strip(),
                    "commit_author": match["commit_author"].strip(),
                    "gerrit": self.get_gerrit_context_of_commit(
                        commit_id, change_id
                    )
                    if self.gerrit_usr
                    else "",
                }
                commits.append(commit)
        return commits

    def get_gerrit_context_of_commit(self, commit_id: str, change_id: str) -> dict:
        if self.gerrit_usr is None:
            return "NO USR"
        if change_id and change_id != "":
            cmd = f"ssh {self.gerrit_usr}@{gerrit_url} -p 29418 gerrit query --format=JSON change:{change_id}"
        elif commit_id and commit_id != "":
            cmd = f"ssh {self.gerrit_usr}@{gerrit_url} -p 29418 gerrit query --format=JSON commit:{commit_id}"
        else:
            logging.warning(
                f"cannot query gerrit, commit_id {commit_id}, change_id {change_id}"
            )
            return "Failed"
        ret, result, err = self.__run_cmd(cmd)
        if ret != 0:
            logging.error("failed to query gerrit for commit %s, ret %d, out: %s, err: %s" % (
                self.commit_id, ret, result, err))
            return None
        gerrit_context = {}
        expected_keys = ("project", "branch", "topic", "url", "subject", "id")
        for line in result.splitlines():
            js = json.loads(line)
            if not js:
                logging.warning(f"failed to load json from str: {line}")
            for key in expected_keys:
                if key in js:
                    gerrit_context[key] = js[key]
        return gerrit_context

    def get_all_branches(self):
        self.local_branch_list = self.get_local_branch_list()
        self.remote_branch_list = self.get_remote_branch_list()
        self.branch_list = self.local_branch_list + self.remote_branch_list
        return self.branch_list

    def get_commit_changes_list_from(self, tag_from: str, tag_to: str) -> list:
        # fetch 远程，获得最新的状态，并且防止找不到
        self.fetch_all_remotes()
        cmd = format("git log --no-merges --pretty=format:\"%%h -%%d %%s (%%cr) <%%an>\" --abbrev-commit %s..%s ." % (tag_from, tag_to))
        ret, result, err = self.__run_cmd(cmd, self.folder)
        if ret != 0:
            logging.error("failed to get remote for folder %s, ret %d, out: %s, err: %s" % (
                self.folder, ret, result, err))
            return None

        self.git_commits = self.__get_git_commit_from_compare_result(result)
        logging.info(f"commits: {pretty_repr(self.git_commits)}")
        return self.git_commits

    def merge_compare_commit_changes_by_changeid(self, commits_forward: list, commits_reverse: list):
        merged_dict_list = []
        change_ids = {}
        # 根据change id去除差异列表里提交相同，但是commit id不同的条目
        for dic in commits_forward:
            dic["operation"] = "++++"
            merged_dict_list.append(dic)
            if "change_id" in dic and len(dic["change_id"]) > 0:
                change_ids[dic["change_id"]] = dic
            elif "id" in dic["gerrit"] and len(dic["gerrit"]["id"] > 0):
                change_ids[dic["gerrit"]["id"]] = dic
            else:
                continue

        for dic in commits_reverse:
            dic["operation"] = "----"
            if "change_id" not in dic and "id" not in dic["gerrit"]:
                merged_dict_list.append(dic)
            elif "change_id" in dic and len(dic["change_id"]) > 0:
                if dic["change_id"] in change_ids:
                    merged_dict_list.remove(change_ids[dic["change_id"]])
                else:
                    merged_dict_list.append(dic)
                    change_ids[dic["change_id"]] = dic
            elif dic["gerrit"]["id"] not in change_ids.keys():
                merged_dict_list.append(dic)
                change_ids[dic["gerrit"]["id"]] = dic
            else:
                continue

        return merged_dict_list

    def compare_between_tags(self, tag_from: str, tag_to: str) -> list:
        git_commits_forward = self.get_commit_changes_list_from(
            tag_from, tag_to)
        git_commits_reverse = self.get_commit_changes_list_from(
            tag_to, tag_from)
        git_commits = self.merge_compare_commit_changes_by_changeid(
            git_commits_forward, git_commits_reverse)
        logging.info(f"merged commits: {pretty_repr(git_commits)}")
        return git_commits


class GitCompare:
    def __init__(self, config_files: list, tag_from: str, tag_to: str, gerrit_usr: str = None) -> None:
        logging.debug(f"init git compare, config files: {config_files}")
        for file in config_files:
            workbook = openpyxl.Workbook()
            folder_list = self.get_concerned_folders(file)
            if folder_list is None:
                continue
            logging.info(f"get folder list: {pretty_repr(folder_list)}")
            self.handle_git_repositiries(
                workbook, folder_list, tag_from, tag_to, gerrit_usr)
            self.set_xml_auto_column_width(workbook)
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            file_name = f'git_compare-{os.path.basename(file)}-{tag_from}-to-{tag_to}.xlsx'
            file_name = self.cleaned_filename(file_name)
            logging.info(f"save result in file: {file_name}")
            workbook.save(file_name)

    def handle_git_repositiries(self, workbook: Workbook, concerned_folders: list, tag_from: str, tag_to: str, gerrit_usr: str = None) -> None:
        bold_font = Font(bold=True)
        for folder_info in concerned_folders:
            if type(folder_info) != dict:
                continue
            git_repos = GitRepository(
                folder_info["name"], folder_info["folder"], gerrit_usr)
            git_commits = git_repos.compare_between_tags(tag_from, tag_to)

            if workbook:
                if folder_info["name"] in workbook.sheetnames:
                    sheet = workbook[folder_info["name"]]
                else:
                    sheet = workbook.create_sheet(folder_info["name"])

                row = 1
                column = 1
                git_infos = {
                    "pwd:": os.path.abspath(git_repos.folder),
                    "local_branch:": git_repos.branch,
                    "remote_branch:": git_repos.remote_branch,
                    "head:": git_repos.head_commit,
                    "date:": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                for key in git_infos:
                    cell = sheet.cell(row=row, column=column)
                    cell.value = key
                    cell.font = bold_font
                    column += 1
                    sheet.merge_cells(format("B%d:G%d" % (row, row)))
                    sheet.cell(row=row, column=column).value = git_infos[key]
                    row += 1
                    column = 1

                row += 2
                column = 1
                sheet.merge_cells(format("A%d:G%d" % (row, row)))
                sheet_info = f"from  {tag_from}  to  {tag_to}"
                cell = sheet.cell(row=row, column=column)
                cell.value = sheet_info
                cell.font = bold_font

                sheet_title = ("ops", "commit-id", "commit-msg", "commit-time", "commit-author", "change-id", "gerrit-url", "gerrit-project", "gerrit-branch")
                row += 1
                for column, title in enumerate(sheet_title, start=1):
                    cell = sheet.cell(row=row, column=column)
                    cell.value = title
                    cell.font = bold_font
                row += 1
                column = 1
                commit_keys = ("operation", "commit_id", "commit_msg", "commit_time", "commit_author", "change_id", "gerrit:url", "gerrit:project", "gerrit:branch")
                for commit in git_commits:
                    for key in commit_keys:
                        if key in commit.keys():
                            sheet.cell(
                                row=row, column=column).value = commit[key]
                        else:
                            sub_keys = key.split(":")
                            if len(sub_keys) == 2 and sub_keys[0] in commit.keys() and type(commit[sub_keys[0]]) == dict and sub_keys[1] in commit[sub_keys[0]].keys():
                                sheet.cell(
                                    row=row, column=column).value = commit[sub_keys[0]][sub_keys[1]]
                        column += 1
                    row += 1
                    column = 1

    def check_if_cell_is_merged(self, ws: Worksheet, cell: Cell) -> bool:
        return any(
            (cell.coordinate in merged_cell)
            for merged_cell in ws.merged_cells.ranges
        )

    def set_xml_auto_column_width(self, wb: Workbook) -> None:
        for ws in wb.worksheets:
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if not cell.value:
                        continue

                    if self.check_if_cell_is_merged(ws, cell):
                        continue
                    # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                    # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                    # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                    cell_len = 0.7 * \
                        len(re.findall('([\u4e00-\u9fa5])',
                            str(cell.value))) + len(str(cell.value))
                    dims[cell.column] = max(
                        (dims.get(cell.column, 0), cell_len))
            for col, value in dims.items():
                # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
                ws.column_dimensions[get_column_letter(col)].width = value + 2

    def cleaned_filename(self, filename) -> str:
        # 定义不允许出现在文件名中的字符
        forbidden_chars = '[\\\\/:*?"<>|]'
        return re.sub(forbidden_chars, '_', filename)

    def get_concerned_folders(self, config_file: str) -> list:
        logging.debug(f"get concerned folders from file: {config_file}")
        if not os.path.exists(config_file):
            logging.error(f"file {config_file} not exist")
            return None
        folder_list = []
        config = configparser.ConfigParser()
        config.read(config_file)
        sections = config.sections()
        for section in sections:
            for key in config[section]:
                folder = config[section][key].rstrip()
                if not os.path.exists(folder):
                    logging.warning(f"folder {folder} not exist")
                    continue
                if not os.path.isdir(folder):
                    logging.warning(f"folder {folder} is not in type of dir")
                    continue
                folder_list.append({"name": key, "folder": folder})

        return folder_list


def main():
    setup_log()
    parser = argparse.ArgumentParser()
    parser.add_argument('-r', '--repo', required=True, nargs=1, type=str, metavar=("<repo_root>"),
                        help='the code root of repo')

    parser.add_argument('-f', '--files', nargs='+', required=True, type=str, metavar=("<files>"),
                        help='the file list concerned folders')

    parser.add_argument('-s', '--tag_from', required=True, nargs=1, type=str, metavar=("<tag|branch|commit>"),
                        help='the compare start tag or branch or commit id')

    parser.add_argument('-t', '--tag_to', required=True, nargs=1, type=str, metavar=("<tag|branch|commit>"),
                        help='the compare end tag or branch or commit id')

    parser.add_argument('-u', '--user', required=False, nargs=1, type=str, metavar=("<gerrit_user_name>"),
                        help='the user name of gerrit')

    args = parser.parse_args()

    if not args.repo:
        logging.error("no repo folder set")
        exit(-1)

    os.chdir(args.repo[0])

    if not args.files:
        logging.error("no concerned folders set")
        exit(-2)
    else:
        GitCompare(
            args.files, args.tag_from[0], args.tag_to[0], args.user[0] if args.user else None)


if __name__ == "__main__":
    gerrit_url = "gerrit.it.xxxx.com"
    ret = 0
    try:
        ret = main()
    except Exception as e:
        logging.error(f'Unexpected error:{str(sys.exc_info()[0])}')
        logging.exception(e)
    sys.exit(ret)
